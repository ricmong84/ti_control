from pathlib import Path
import re
from zipfile import BadZipFile
from openpyxl import load_workbook

from app import app
from models.models import db, ActivoTI

EXCEL_PATH = Path(__file__).with_name("HW_activos_limpio.xlsx")

MISSING_MARKERS = {
    "", "-", "--", "---", "----", "-----",
    ".", "..", "...", ",", ".,", ".-.", ".-.,",
    "\\---", "\\--",
    "N/A", "NO INDICA", "NO TIENE", "NO TIENE 060", "FF"
}


def normalize(value):
    if value is None:
        return None

    if isinstance(value, bool):
        return str(value)

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value).strip()

    text = str(value).replace("\u200b", "").strip()
    text = re.sub(r"\s+", " ", text).strip()

    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]

    return text or None


def is_missing(text):
    if text is None:
        return True

    upper = text.upper().strip()

    if upper in MISSING_MARKERS:
        return True

    if "FALTA SERIE" in upper:
        return True

    if re.fullmatch(r"[-.,\\\/ ]+", text):
        return True

    return False


def clean_placa(placa):
    placa = normalize(placa)
    if placa is None:
        return None

    upper = placa.upper()

    if upper.startswith("PLACA SIFA "):
        placa = placa.split(" ", 2)[-1].strip()

    placa = re.sub(r"\s+", "", placa)
    return placa or None


def main():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"No se encontró el archivo Excel: {EXCEL_PATH}")

    try:
        wb = load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    except BadZipFile:
        raise Exception(
            f"El archivo '{EXCEL_PATH.name}' no es un .xlsx válido.\n"
            "Abre el archivo en Excel y vuelve a guardarlo como 'Libro de Excel (*.xlsx)'."
        )

    ws = wb[wb.sheetnames[0]]

    created = 0
    updated = 0
    skipped = 0

    with app.app_context():
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            tipo, marca, modelo, serie, placa = (tuple(row) + (None,) * 5)[:5]

            tipo = normalize(tipo)
            marca = normalize(marca)
            modelo = normalize(modelo)
            serie = normalize(serie)
            placa = clean_placa(placa)

            if is_missing(placa):
                skipped += 1
                continue

            if is_missing(serie):
                serie = None
            if is_missing(tipo):
                tipo = None
            if is_missing(marca):
                marca = None
            if is_missing(modelo):
                modelo = None

            row_db = ActivoTI.query.filter_by(placa_sifa=placa).first()

            if not row_db:
                row_db = ActivoTI(
                    placa_sifa=placa,
                    serie_service_tag=serie,
                    marca=marca,
                    modelo=modelo,
                    tipo_equipo=tipo,
                    ip=None,
                    activo=True
                )
                db.session.add(row_db)
                created += 1
            else:
                row_db.serie_service_tag = serie
                row_db.marca = marca
                row_db.modelo = modelo
                row_db.tipo_equipo = tipo
                updated += 1

            if (created + updated) % 1000 == 0:
                db.session.flush()

        db.session.commit()

        total = ActivoTI.query.count()
        print(
            f"✅ Seed Activos completado. "
            f"Nuevos: {created} | "
            f"Actualizados: {updated} | "
            f"Omitidos sin placa: {skipped} | "
            f"Total BD: {total}"
        )


if __name__ == "__main__":
    main()