from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from io import BytesIO
from datetime import datetime

from models.models import db, SuministroImpresora, SuministroMovimiento
from routes.utils import admin_required

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

inventario_bp = Blueprint("inventario", __name__, url_prefix="/inventario")


@inventario_bp.route("/")
@login_required
def lista():
    q = request.args.get("q", "").strip()

    query = SuministroImpresora.query
    if q:
        like = f"%{q}%"
        query = query.filter(
            (SuministroImpresora.modelo.ilike(like)) |
            (SuministroImpresora.consumible.ilike(like))
        )

    items = query.order_by(SuministroImpresora.modelo.asc(), SuministroImpresora.consumible.asc()).all()

    has_low = any(i.actual < i.minimo for i in items)

    grouped = {}
    for it in items:
        grouped.setdefault(it.modelo, []).append(it)

    return render_template("inventario/list.html", grouped=grouped, q=q, has_low=has_low)


# ✅ Ajuste de ACTUAL (+/-) -> lo pueden usar admin y user
@inventario_bp.route("/<int:item_id>/adjust", methods=["POST"])
@login_required
def adjust(item_id):
    data = request.get_json(silent=True) or {}
    delta = int(data.get("delta", 0))

    item = SuministroImpresora.query.get_or_404(item_id)
    new_val = item.actual + delta
    if new_val < 0:
        new_val = 0

    item.actual = new_val

    mov = SuministroMovimiento(
        suministro_id=item.id,
        user_id=current_user.id,
        delta=delta
    )
    db.session.add(mov)
    db.session.commit()

    return jsonify({
        "ok": True,
        "id": item.id,
        "actual": item.actual,
        "minimo": item.minimo,
        "is_low": item.actual < item.minimo
    })


# ✅ Crear suministro (SOLO ADMIN)
@inventario_bp.route("/create", methods=["GET", "POST"])
@login_required
@admin_required
def create():
    if request.method == "POST":
        modelo = request.form.get("modelo", "").strip()
        consumible = request.form.get("consumible", "").strip()

        # ✅ AQUÍ VA EL BLOQUE QUE PREGUNTABAS
        actual = max(int(request.form.get("actual", 0)), 0)
        minimo = max(int(request.form.get("minimo", 0)), 0)
        maximo_raw = request.form.get("maximo", "").strip()
        maximo = int(maximo_raw) if maximo_raw else None

        if not modelo or not consumible:
            flash("Modelo y Consumible son obligatorios.", "danger")
            return render_template("inventario/form.html", mode="create")

        exists = SuministroImpresora.query.filter_by(modelo=modelo, consumible=consumible).first()
        if exists:
            flash("Ese suministro ya existe para ese modelo.", "warning")
            return render_template("inventario/form.html", mode="create")

        row = SuministroImpresora(
            modelo=modelo,
            consumible=consumible,
            actual=actual,
            minimo=minimo,
            maximo=maximo
        )
        db.session.add(row)
        db.session.commit()

        flash("Suministro agregado.", "success")
        return redirect(url_for("inventario.lista"))

    return render_template("inventario/form.html", mode="create")


# ✅ Editar suministro (SOLO ADMIN) -> admin puede editar actual/min/max
@inventario_bp.route("/<int:item_id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def edit(item_id):
    item = SuministroImpresora.query.get_or_404(item_id)

    if request.method == "POST":
        item.modelo = request.form.get("modelo", "").strip()
        item.consumible = request.form.get("consumible", "").strip()

        # ✅ AQUÍ TAMBIÉN VA EL BLOQUE (con defaults del item)
        item.actual = max(int(request.form.get("actual", item.actual)), 0)
        item.minimo = max(int(request.form.get("minimo", item.minimo)), 0)
        maximo_raw = request.form.get("maximo", "").strip()
        item.maximo = int(maximo_raw) if maximo_raw else None

        db.session.commit()
        flash("Suministro actualizado.", "success")
        return redirect(url_for("inventario.lista"))

    return render_template("inventario/form.html", mode="edit", item=item)


# ✅ Eliminar suministro (SOLO ADMIN)
@inventario_bp.route("/<int:item_id>/delete", methods=["POST"])
@login_required
@admin_required
def delete(item_id):
    item = SuministroImpresora.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash("Suministro eliminado.", "info")
    return redirect(url_for("inventario.lista"))


# ✅ Exportar solo los que están bajo mínimo
@inventario_bp.route("/export_bajo_minimo", methods=["GET"])
@login_required
def export_bajo_minimo():
    low = SuministroImpresora.query.filter(SuministroImpresora.actual < SuministroImpresora.minimo) \
                                  .order_by(SuministroImpresora.modelo.asc(), SuministroImpresora.consumible.asc()) \
                                  .all()

    if not low:
        flash("No hay suministros por debajo del mínimo.", "info")
        return redirect(url_for("inventario.lista"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Bajo_Minimo"

    headers = ["Modelo", "Consumible", "Actual", "Mínimo", "Máximo"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for it in low:
        ws.append([it.modelo, it.consumible, it.actual, it.minimo, (it.maximo if it.maximo is not None else "")])

    widths = {1: 14, 2: 28, 3: 10, 4: 10, 5: 10}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"inventario_bajo_minimo_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )