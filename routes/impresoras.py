from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from datetime import datetime
from io import BytesIO

from models.models import db, Impresora, ReporteImpresora
from routes.utils import admin_required

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

impresoras_bp = Blueprint("impresoras", __name__, url_prefix="/impresoras")


# -------------------------
# Helpers
# -------------------------
def can_access_report(r: ReporteImpresora) -> bool:
    return current_user.role == "admin" or r.user_id == current_user.id


def _safe_sheet_title(name: str) -> str:
    """Excel: máximo 31 caracteres y sin símbolos inválidos."""
    forbidden = ['\\', '/', '*', '[', ']', ':', '?']
    for ch in forbidden:
        name = name.replace(ch, '-')
    return name[:31]


def _write_averias_sheet(ws, r: ReporteImpresora):
    """
    Genera el formato tipo 'ficha' como tu imagen (Averías):
    2 columnas (Etiqueta / Valor), colores oscuros y bordes azules.
    """
    # Colores
    bg_header = PatternFill("solid", fgColor="1F2933")  # oscuro
    bg_label  = PatternFill("solid", fgColor="3B4A5A")  # gris-azulado
    bg_value  = PatternFill("solid", fgColor="2A2F36")  # gris oscuro
    border_c  = "6EA8FE"                                # azul claro

    thin = Side(style="thin", color=border_c)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    font_title = Font(bold=True, size=18, color="FFFFFF")
    font_label = Font(bold=True, size=11, color="FFFFFF")
    font_value = Font(size=11, color="FFFFFF")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Tamaños
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 52

    # Título
    ws.merge_cells("A1:B1")
    ws["A1"].value = "Reportes de Averías"
    ws["A1"].fill = bg_header
    ws["A1"].font = font_title
    ws["A1"].alignment = align_center
    ws["A1"].border = border
    ws["B1"].border = border
    ws.row_dimensions[1].height = 32

    # Sucursal / Departamento (derivado de ubicación)
    sucursal = "HDT"
    if r.ubicacion and "-" in r.ubicacion:
        first = r.ubicacion.split("-")[0].strip()
        if first:
            sucursal = first

    telefono = r.extension if getattr(r, "extension", None) else "Teams"

    rows = [
        ("CLIENTE:", r.contacto),
        ("UBICACIÓN (Sucursal / Departamento):", sucursal),
        ("DIRECCIÓN EXACTA:", r.ubicacion),
        ("MODELO:", r.modelo),
        ("NÚMERO DE SERIE:", r.serie),
        ("DESCRIPCIÓN DEL PROBLEMA:\nIndicar código detallado en pantalla", r.descripcion),
        ("CONTACTO:", r.contacto),
        ("TELÉFONO:", telefono),
        ("IP", r.ip),
    ]

    start_row = 2
    for i, (label, value) in enumerate(rows):
        row = start_row + i
        ws.row_dimensions[row].height = 28 if i != 5 else 44  # descripción más alta

        a = ws.cell(row=row, column=1, value=label)
        b = ws.cell(row=row, column=2, value=value)

        a.fill = bg_label
        a.font = font_label
        a.alignment = align_center
        a.border = border

        b.fill = bg_value
        b.font = font_value
        b.alignment = align_center
        b.border = border

    # Ajustes impresión (opcional)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True


# =========================================================
# A) CATÁLOGO IMPRESORAS (SOLO ADMIN)
# =========================================================
@impresoras_bp.route("/catalogo")
@login_required
@admin_required
def catalogo_list():
    q = request.args.get("q", "").strip()
    query = Impresora.query
    if q:
        like = f"%{q}%"
        query = query.filter(
            (Impresora.ip.ilike(like)) |
            (Impresora.serie.ilike(like)) |
            (Impresora.modelo.ilike(like)) |
            (Impresora.ubicacion.ilike(like)) |
            (Impresora.servidor.ilike(like))
        )
    impresoras = query.order_by(Impresora.ip.asc()).all()
    return render_template("impresoras/catalogo_list.html", impresoras=impresoras, q=q)


@impresoras_bp.route("/catalogo/create", methods=["GET", "POST"])
@login_required
@admin_required
def catalogo_create():
    if request.method == "POST":
        ip = request.form.get("ip", "").strip()
        serie = request.form.get("serie", "").strip()
        modelo = request.form.get("modelo", "").strip()
        ubicacion = request.form.get("ubicacion", "").strip()
        servidor = request.form.get("servidor", "").strip()

        if not (ip and serie and modelo and ubicacion and servidor):
            flash("Todos los campos son obligatorios.", "danger")
            return render_template("impresoras/catalogo_form.html", mode="create")

        if Impresora.query.filter_by(ip=ip).first():
            flash("Ya existe una impresora con esa IP.", "warning")
            return render_template("impresoras/catalogo_form.html", mode="create")

        row = Impresora(ip=ip, serie=serie, modelo=modelo, ubicacion=ubicacion, servidor=servidor)
        db.session.add(row)
        db.session.commit()
        flash("Impresora creada.", "success")
        return redirect(url_for("impresoras.catalogo_list"))

    return render_template("impresoras/catalogo_form.html", mode="create")


@impresoras_bp.route("/catalogo/<int:imp_id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def catalogo_edit(imp_id):
    imp = Impresora.query.get_or_404(imp_id)

    if request.method == "POST":
        ip = request.form.get("ip", "").strip()
        serie = request.form.get("serie", "").strip()
        modelo = request.form.get("modelo", "").strip()
        ubicacion = request.form.get("ubicacion", "").strip()
        servidor = request.form.get("servidor", "").strip()

        if not (ip and serie and modelo and ubicacion and servidor):
            flash("Todos los campos son obligatorios.", "danger")
            return render_template("impresoras/catalogo_form.html", mode="edit", imp=imp)

        exists = Impresora.query.filter_by(ip=ip).first()
        if exists and exists.id != imp.id:
            flash("Esa IP ya está en uso.", "warning")
            return render_template("impresoras/catalogo_form.html", mode="edit", imp=imp)

        imp.ip = ip
        imp.serie = serie
        imp.modelo = modelo
        imp.ubicacion = ubicacion
        imp.servidor = servidor

        db.session.commit()
        flash("Impresora actualizada.", "success")
        return redirect(url_for("impresoras.catalogo_list"))

    return render_template("impresoras/catalogo_form.html", mode="edit", imp=imp)


@impresoras_bp.route("/catalogo/<int:imp_id>/delete", methods=["POST"])
@login_required
@admin_required
def catalogo_delete(imp_id):
    imp = Impresora.query.get_or_404(imp_id)
    db.session.delete(imp)
    db.session.commit()
    flash("Impresora eliminada.", "info")
    return redirect(url_for("impresoras.catalogo_list"))


# =========================================================
# B) REPORTES (ADMIN: TODOS / USER: SOLO SUYOS)
# =========================================================
from sqlalchemy import or_

@impresoras_bp.route("/reportes")
@login_required
def reportes_list():
    q = request.args.get("q", "").strip()

    if current_user.role == "admin":
        query = ReporteImpresora.query
    else:
        query = ReporteImpresora.query.filter_by(user_id=current_user.id)

    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            ReporteImpresora.ip.ilike(like),
            ReporteImpresora.modelo.ilike(like),
            ReporteImpresora.contacto.ilike(like),
            ReporteImpresora.ubicacion.ilike(like)
        ))

    reportes = query.order_by(ReporteImpresora.id.desc()).all()
    return render_template("impresoras/reportes_list.html", reportes=reportes, q=q)


@impresoras_bp.route("/reportes/create", methods=["GET", "POST"])
@login_required
def reportes_create():
    impresoras = Impresora.query.order_by(Impresora.ip.asc()).all()

    if request.method == "POST":
        impresora_id = int(request.form.get("impresora_id"))
        imp = Impresora.query.get_or_404(impresora_id)

        contacto = request.form.get("contacto", "").strip()
        extension = request.form.get("extension", "").strip()
        descripcion = request.form.get("descripcion", "").strip()

        if not contacto or not descripcion:
            flash("Contacto y descripción son obligatorios.", "danger")
            return render_template("impresoras/form.html", impresoras=impresoras, mode="create")

        r = ReporteImpresora(
            user_id=current_user.id,
            impresora_id=imp.id,
            ip=imp.ip,
            serie=imp.serie,
            modelo=imp.modelo,
            ubicacion=imp.ubicacion,
            contacto=contacto,
            extension=extension if extension else None,
            descripcion=descripcion
        )
        db.session.add(r)
        db.session.commit()
        flash("Reporte creado.", "success")
        return redirect(url_for("impresoras.reportes_list"))

    return render_template("impresoras/form.html", impresoras=impresoras, mode="create")


@impresoras_bp.route("/reportes/<int:rep_id>/edit", methods=["GET", "POST"])
@login_required
def reportes_edit(rep_id):
    rep = ReporteImpresora.query.get_or_404(rep_id)
    if not can_access_report(rep):
        return "Acceso denegado", 403

    impresoras = Impresora.query.order_by(Impresora.ip.asc()).all()

    if request.method == "POST":
        impresora_id = int(request.form.get("impresora_id"))
        imp = Impresora.query.get_or_404(impresora_id)

        rep.impresora_id = imp.id
        rep.ip = imp.ip
        rep.serie = imp.serie
        rep.modelo = imp.modelo
        rep.ubicacion = imp.ubicacion

        rep.contacto = request.form.get("contacto", "").strip()
        ext = request.form.get("extension", "").strip()
        rep.extension = ext if ext else None
        rep.descripcion = request.form.get("descripcion", "").strip()

        db.session.commit()
        flash("Reporte actualizado.", "success")
        return redirect(url_for("impresoras.reportes_list"))

    return render_template("impresoras/form.html", impresoras=impresoras, reporte=rep, mode="edit")


@impresoras_bp.route("/reportes/<int:rep_id>/delete", methods=["POST"])
@login_required
def reportes_delete(rep_id):
    rep = ReporteImpresora.query.get_or_404(rep_id)
    if not can_access_report(rep):
        return "Acceso denegado", 403

    db.session.delete(rep)
    db.session.commit()
    flash("Reporte eliminado.", "info")
    return redirect(url_for("impresoras.reportes_list"))


# API para autollenado por IP (select)
@impresoras_bp.route("/api/<int:imp_id>")
@login_required
def api_impresora(imp_id):
    i = Impresora.query.get_or_404(imp_id)
    return jsonify({
        "serie": i.serie,
        "modelo": i.modelo,
        "ubicacion": i.ubicacion,
        "servidor": i.servidor
    })


# =========================================================
# EXPORT EXCEL (FORMATO AVERÍAS)
# =========================================================
@impresoras_bp.route("/reportes/export", methods=["POST"])
@login_required
def reportes_export():
    export_all = request.form.get("export_all") == "1"
    selected_ids = request.form.getlist("selected_ids")

    if current_user.role == "admin":
        query = ReporteImpresora.query
    else:
        query = ReporteImpresora.query.filter_by(user_id=current_user.id)

    if not export_all:
        if not selected_ids:
            flash("Debes seleccionar al menos un reporte para exportar.", "warning")
            return redirect(url_for("impresoras.reportes_list"))
        ids_int = [int(x) for x in selected_ids]
        query = query.filter(ReporteImpresora.id.in_(ids_int))

    rows = query.order_by(ReporteImpresora.id.asc()).all()

    wb = Workbook()
    wb.remove(wb.active)  # eliminar hoja por defecto

    for r in rows:
        title = _safe_sheet_title(f"Averia_{r.id}")
        ws = wb.create_sheet(title=title)
        _write_averias_sheet(ws, r)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reportes_averias_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )