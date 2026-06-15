from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
from sqlalchemy import or_
from datetime import datetime

from models.models import db, ActivoTI, ReporteEquipo
from routes.utils import admin_required
from flask import send_file
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

activos_bp = Blueprint("activos", __name__, url_prefix="/activos")


# =========================================================
# A) CATÁLOGO DE ACTIVOS (SOLO ADMIN)
# =========================================================
@activos_bp.route("/")
@login_required
@admin_required
def activos_list():
    q = request.args.get("q", "").strip()
    query = ActivoTI.query

    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            ActivoTI.placa_sifa.ilike(like),
            ActivoTI.serie_service_tag.ilike(like),
            ActivoTI.marca.ilike(like),
            ActivoTI.modelo.ilike(like),
            ActivoTI.tipo_equipo.ilike(like),
            ActivoTI.ip.ilike(like)
        ))

    activos = query.order_by(ActivoTI.placa_sifa.asc()).all()
    return render_template("activos/list.html", activos=activos, q=q)


@activos_bp.route("/create", methods=["GET", "POST"])
@login_required
@admin_required
def activos_create():
    if request.method == "POST":
        placa_sifa = request.form.get("placa_sifa", "").strip()
        serie_service_tag = request.form.get("serie_service_tag", "").strip()
        marca = request.form.get("marca", "").strip()
        modelo = request.form.get("modelo", "").strip()
        tipo_equipo = request.form.get("tipo_equipo", "").strip()
        ip = request.form.get("ip", "").strip()

        if not placa_sifa:
            flash("La placa SIFA es obligatoria.", "danger")
            return render_template("activos/form.html", mode="create")

        exists = ActivoTI.query.filter_by(placa_sifa=placa_sifa).first()
        if exists:
            flash("Ya existe un activo con esa placa SIFA.", "warning")
            return render_template("activos/form.html", mode="create")

        row = ActivoTI(
            placa_sifa=placa_sifa,
            serie_service_tag=serie_service_tag or None,
            marca=marca or None,
            modelo=modelo or None,
            tipo_equipo=tipo_equipo or None,
            ip=ip or None
        )
        db.session.add(row)
        db.session.commit()
        flash("Activo creado correctamente.", "success")
        return redirect(url_for("activos.activos_list"))

    return render_template("activos/form.html", mode="create")


@activos_bp.route("/<int:activo_id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def activos_edit(activo_id):
    activo = ActivoTI.query.get_or_404(activo_id)

    if request.method == "POST":
        placa_sifa = request.form.get("placa_sifa", "").strip()
        serie_service_tag = request.form.get("serie_service_tag", "").strip()
        marca = request.form.get("marca", "").strip()
        modelo = request.form.get("modelo", "").strip()
        tipo_equipo = request.form.get("tipo_equipo", "").strip()
        ip = request.form.get("ip", "").strip()

        if not placa_sifa:
            flash("La placa SIFA es obligatoria.", "danger")
            return render_template("activos/form.html", mode="edit", activo=activo)

        exists = ActivoTI.query.filter_by(placa_sifa=placa_sifa).first()
        if exists and exists.id != activo.id:
            flash("Esa placa SIFA ya está registrada.", "warning")
            return render_template("activos/form.html", mode="edit", activo=activo)

        activo.placa_sifa = placa_sifa
        activo.serie_service_tag = serie_service_tag or None
        activo.marca = marca or None
        activo.modelo = modelo or None
        activo.tipo_equipo = tipo_equipo or None
        activo.ip = ip or None

        db.session.commit()
        flash("Activo actualizado.", "success")
        return redirect(url_for("activos.activos_list"))

    return render_template("activos/form.html", mode="edit", activo=activo)


@activos_bp.route("/<int:activo_id>/delete", methods=["POST"])
@login_required
@admin_required
def activos_delete(activo_id):
    activo = ActivoTI.query.get_or_404(activo_id)
    db.session.delete(activo)
    db.session.commit()
    flash("Activo eliminado.", "info")
    return redirect(url_for("activos.activos_list"))


# API para autocompletar por placa SIFA
@activos_bp.route("/api/placa/<placa_sifa>")
@login_required
def api_placa(placa_sifa):
    activo = ActivoTI.query.filter_by(placa_sifa=placa_sifa).first()
    if not activo:
        return jsonify({"ok": False})

    return jsonify({
        "ok": True,
        "id": activo.id,
        "serie_service_tag": activo.serie_service_tag or "",
        "marca": activo.marca or "",
        "modelo": activo.modelo or "",
        "tipo_equipo": activo.tipo_equipo or "",
        "ip": activo.ip or ""
    })


# =========================================================
# B) REPORTES DE EQUIPO (TODOS VEN TODOS)
# =========================================================
@activos_bp.route("/reportes")
@login_required
def reportes_list():
    q = request.args.get("q", "").strip()
    query = ReporteEquipo.query

    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            ReporteEquipo.placa_sifa.ilike(like),
            ReporteEquipo.serie_service_tag.ilike(like),
            ReporteEquipo.marca.ilike(like),
            ReporteEquipo.modelo.ilike(like),
            ReporteEquipo.tipo_equipo.ilike(like),
            ReporteEquipo.usuario_reporta.ilike(like),
            ReporteEquipo.tecnico.ilike(like),
            ReporteEquipo.edificio.ilike(like),
            ReporteEquipo.sas.ilike(like),
            ReporteEquipo.observaciones.ilike(like),
            ReporteEquipo.extension.ilike(like),
            ReporteEquipo.direccion_ip.ilike(like),
            ReporteEquipo.ubicacion.ilike(like),
            ReporteEquipo.estado.ilike(like)
            
        ))

    reportes = query.order_by(ReporteEquipo.id.desc()).all()
    activos = ActivoTI.query.order_by(ActivoTI.placa_sifa.asc()).all()

    return render_template("activos/reportes_list.html", reportes=reportes, q=q, activos=activos)


@activos_bp.route("/reportes/create", methods=["GET", "POST"])
@login_required
def reportes_create():
    activos = ActivoTI.query.order_by(ActivoTI.placa_sifa.asc()).all()

    if request.method == "POST":
        placa_sifa = request.form.get("placa_sifa", "").strip()
        serie_service_tag = request.form.get("serie_service_tag", "").strip()
        marca = request.form.get("marca", "").strip()
        modelo = request.form.get("modelo", "").strip()
        tipo_equipo = request.form.get("tipo_equipo", "").strip()
        direccion_ip = request.form.get("direccion_ip", "").strip()
        usuario_reporta = request.form.get("usuario_reporta", "").strip()
        tecnico = request.form.get("tecnico", "").strip()
        observaciones = request.form.get("observaciones", "").strip()
        extension = request.form.get("extension", "").strip()
        ubicacion = request.form.get("ubicacion", "").strip()
        edificio = request.form.get("edificio", "").strip()
        sas = request.form.get("sas", "").strip()

        if not usuario_reporta or not observaciones:
            flash("Usuario que reporta y observaciones son obligatorios.", "danger")
            return render_template("activos/reporte_form.html", mode="create", activos=activos)

        if not placa_sifa and not direccion_ip:
            flash("Debes indicar una Placa SIFA o una Dirección IP.", "warning")
            return render_template("activos/reporte_form.html", mode="create", activos=activos)

        if placa_sifa and direccion_ip:
            direccion_ip = None  # por regla, si tiene placa, no se usa IP

        activo = ActivoTI.query.filter_by(placa_sifa=placa_sifa).first() if placa_sifa else None

        reporte = ReporteEquipo(
        activo_id=activo.id if activo else None,
        placa_sifa=placa_sifa or None,
        serie_service_tag=serie_service_tag or None,
        marca=marca or None,
        modelo=modelo or None,
        tipo_equipo=tipo_equipo or None,
        direccion_ip=direccion_ip or None,
        ubicacion=ubicacion or None,
        edificio=edificio or None,
        sas=sas or None,
        usuario_reporta=usuario_reporta,
        tecnico=tecnico or None,
        observaciones=observaciones,
        extension=extension or None
)

        db.session.add(reporte)
        db.session.commit()
        flash("Reporte creado correctamente.", "success")
        return redirect(url_for("activos.reportes_list"))

    return render_template("activos/reporte_form.html", mode="create", activos=activos)


@activos_bp.route("/reportes/<int:reporte_id>/edit", methods=["GET", "POST"])
@login_required
def reportes_edit(reporte_id):
    reporte = ReporteEquipo.query.get_or_404(reporte_id)
    activos = ActivoTI.query.order_by(ActivoTI.placa_sifa.asc()).all()

    if request.method == "POST":
        placa_sifa = request.form.get("placa_sifa", "").strip()
        serie_service_tag = request.form.get("serie_service_tag", "").strip()
        marca = request.form.get("marca", "").strip()
        modelo = request.form.get("modelo", "").strip()
        tipo_equipo = request.form.get("tipo_equipo", "").strip()
        direccion_ip = request.form.get("direccion_ip", "").strip()
        usuario_reporta = request.form.get("usuario_reporta", "").strip()
        tecnico = request.form.get("tecnico", "").strip()
        observaciones = request.form.get("observaciones", "").strip()
        extension = request.form.get("extension", "").strip()
        estado = request.form.get("estado", "pendiente").strip()
        ubicacion = request.form.get("ubicacion", "").strip()
        edificio = request.form.get("edificio", "").strip()
        sas = request.form.get("sas", "").strip()


        if not usuario_reporta or not observaciones:
            flash("Usuario que reporta y observaciones son obligatorios.", "danger")
            return render_template("activos/reporte_form.html", mode="edit", activos=activos, reporte=reporte)

        if not placa_sifa and not direccion_ip:
            flash("Debes indicar una Placa SIFA o una Dirección IP.", "warning")
            return render_template("activos/reporte_form.html", mode="edit", activos=activos, reporte=reporte)

        if placa_sifa and direccion_ip:
            direccion_ip = None

        activo = ActivoTI.query.filter_by(placa_sifa=placa_sifa).first() if placa_sifa else None

        reporte.activo_id = activo.id if activo else None
        reporte.placa_sifa = placa_sifa or None
        reporte.serie_service_tag = serie_service_tag or None
        reporte.marca = marca or None
        reporte.modelo = modelo or None
        reporte.tipo_equipo = tipo_equipo or None
        reporte.direccion_ip = direccion_ip or None
        reporte.usuario_reporta = usuario_reporta
        reporte.tecnico = tecnico or None
        reporte.observaciones = observaciones
        reporte.extension = extension or None
        reporte.estado = estado
        reporte.ubicacion = ubicacion or None
        reporte.edificio = edificio or None
        reporte.sas = sas or None


        if estado == "atendido" and not reporte.fecha_atencion:
            reporte.fecha_atencion = datetime.utcnow()
        elif estado != "atendido":
            reporte.fecha_atencion = None

        db.session.commit()
        flash("Reporte actualizado.", "success")
        return redirect(url_for("activos.reportes_list"))

    return render_template("activos/reporte_form.html", mode="edit", activos=activos, reporte=reporte)


@activos_bp.route("/reportes/<int:reporte_id>/atender", methods=["POST"])
@login_required
def reportes_atender(reporte_id):
    reporte = ReporteEquipo.query.get_or_404(reporte_id)
    reporte.estado = "atendido"
    reporte.fecha_atencion = datetime.utcnow()
    db.session.commit()
    flash("Reporte marcado como atendido.", "success")
    return redirect(url_for("activos.reportes_list"))

@activos_bp.route("/reportes/export", methods=["POST"])
@login_required
def reportes_export():
    selected_ids = request.form.getlist("selected_ids")

    if not selected_ids:
        flash("Debes seleccionar al menos un reporte para exportar.", "warning")
        return redirect(url_for("activos.reportes_list"))

    ids_int = [int(x) for x in selected_ids]
    rows = ReporteEquipo.query.filter(ReporteEquipo.id.in_(ids_int)) \
                              .order_by(ReporteEquipo.id.asc()) \
                              .all()

    wb = Workbook()
    wb.remove(wb.active)

    for r in rows:
        title = _safe_sheet_title(f"Reporte_{r.id}")
        ws = wb.create_sheet(title=title)
        _write_reporte_equipo_sheet(ws, r)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reportes_equipo_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@activos_bp.route("/reportes/<int:reporte_id>/delete", methods=["POST"])
@login_required
@admin_required
def reportes_delete(reporte_id):
    reporte = ReporteEquipo.query.get_or_404(reporte_id)
    db.session.delete(reporte)
    db.session.commit()
    flash("Reporte eliminado.", "info")
    return redirect(url_for("activos.reportes_list"))

def _safe_sheet_title(name: str) -> str:
    forbidden = ['\\', '/', '*', '[', ']', ':', '?']
    for ch in forbidden:
        name = name.replace(ch, '-')
    return name[:31]


def _write_reporte_equipo_sheet(ws, r):
    """
    Formato Excel tipo ficha según imagen:
    5 columnas:
    A Tipo de equipo
    B Marca
    C Serie
    D Service Tag
    E Placa Sifa
    """

    # Colores
    fill_header = PatternFill("solid", fgColor="D9D9D9")   # gris claro
    fill_value = PatternFill("solid", fgColor="FFF200")    # amarillo
    fill_gray = PatternFill("solid", fgColor="A6A6A6")     # gris medio

    bold = Font(bold=True, size=12, color="000000")
    normal = Font(size=11, color="000000")

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Anchos de columnas
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    # Alturas
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 24
    ws.row_dimensions[6].height = 30
    ws.row_dimensions[7].height = 24
    ws.row_dimensions[8].height = 40
    ws.row_dimensions[9].height = 40
    ws.row_dimensions[10].height = 24
    ws.row_dimensions[11].height = 24

    # --------------------------
    # Fila 1: encabezados
    # --------------------------
    headers = ["Tipo de Equipo", "Marca", "Serie", "Service Tag", "Placa Sifa"]
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=text)
        c.fill = fill_header
        c.font = bold
        c.alignment = center
        c.border = border

    # --------------------------
    # Fila 2: valores
    # --------------------------
    # Asunción:
    # - Serie queda vacía porque tu modelo actual solo tiene serie_service_tag
    # - Service Tag usa serie_service_tag
    values = [
        r.tipo_equipo or "",
        r.marca or "",
        "",  # Serie
        r.serie_service_tag or "",
        r.placa_sifa or ""
    ]

    for col, text in enumerate(values, start=1):
        c = ws.cell(row=2, column=col, value=text)
        c.fill = fill_gray if col != 4 else fill_value
        c.font = normal
        c.alignment = center
        c.border = border

    # --------------------------
    # Fila 3: Usuario / Extensión
    # --------------------------
    ws.merge_cells("A3:D3")
    c = ws["A3"]
    c.value = "Usuario"
    c.fill = fill_value
    c.font = bold
    c.alignment = center
    c.border = border

    for cell in ["B3", "C3", "D3"]:
        ws[cell].border = border

    c = ws["E3"]
    c.value = "Extensión"
    c.fill = fill_header
    c.font = bold
    c.alignment = center
    c.border = border

    # --------------------------
    # Fila 4: valor usuario / extensión
    # --------------------------
    ws.merge_cells("A4:D4")
    c = ws["A4"]
    c.value = r.usuario_reporta or ""
    c.fill = fill_gray
    c.font = normal
    c.alignment = center
    c.border = border

    for cell in ["B4", "C4", "D4"]:
        ws[cell].border = border

    c = ws["E4"]
    c.value = r.extension or ""
    c.fill = fill_value
    c.font = normal
    c.alignment = center
    c.border = border

    # --------------------------
    # Fila 5: Edificio / Ubicación exacta
    # --------------------------
    c = ws["A5"]
    c.value = "Edificio"
    c.fill = fill_header
    c.font = bold
    c.alignment = center
    c.border = border

    ws.merge_cells("B5:E5")
    c = ws["B5"]
    c.value = "Ubicación exacta"
    c.fill = fill_header
    c.font = bold
    c.alignment = center
    c.border = border

    for cell in ["C5", "D5", "E5"]:
        ws[cell].border = border

    # --------------------------
    # Fila 6: valores Edificio / Ubicación
    # --------------------------
    # Asunción:
    # - Edificio se deja en blanco
    # - Ubicación usa r.ubicacion
    c = ws["A6"]
    c.value = r.edificio or ""
    c.fill = fill_gray
    c.font = normal
    c.alignment = center
    c.border = border

    ws.merge_cells("B6:E6")
    c = ws["B6"]
    c.value = r.ubicacion or ""
    c.fill = fill_value
    c.font = normal
    c.alignment = center
    c.border = border

    for cell in ["C6", "D6", "E6"]:
        ws[cell].border = border

    # --------------------------
    # Fila 7: Diagnóstico u observaciones
    # --------------------------
    ws.merge_cells("A7:E7")
    c = ws["A7"]
    c.value = "Diagnóstico u observaciones"
    c.fill = fill_header
    c.font = bold
    c.alignment = center
    c.border = border
    for cell in ["B7", "C7", "D7", "E7"]:
        ws[cell].border = border

    # --------------------------
    # Fila 8-9: observaciones grandes
    # --------------------------
    ws.merge_cells("A8:E9")
    c = ws["A8"]
    c.value = r.observaciones or ""
    c.fill = fill_value
    c.font = normal
    c.alignment = left
    c.border = border
    for cell in ["B8", "C8", "D8", "E8", "A9", "B9", "C9", "D9", "E9"]:
        ws[cell].border = border

    # --------------------------
    # Fila 10: Técnico / SAS / Fecha
    # --------------------------
    ws.merge_cells("A10:C10")
    c = ws["A10"]
    c.value = "Técnico"
    c.fill = fill_header
    c.font = bold
    c.alignment = center
    c.border = border
    for cell in ["B10", "C10"]:
        ws[cell].border = border

    c = ws["D10"]
    c.value = "SAS"
    c.fill = fill_header
    c.font = bold
    c.alignment = center
    c.border = border

    c = ws["E10"]
    c.value = "Fecha"
    c.fill = fill_header
    c.font = bold
    c.alignment = center
    c.border = border

    # --------------------------
    # Fila 11: valores Técnico / SAS / Fecha
    # --------------------------
    ws.merge_cells("A11:C11")
    c = ws["A11"]
    c.value = r.tecnico or ""
    c.fill = fill_gray
    c.font = normal
    c.alignment = center
    c.border = border
    for cell in ["B11", "C11"]:
        ws[cell].border = border

    c = ws["D11"]
    c.value = r.sas or ""
    c.fill = fill_value
    c.font = normal
    c.alignment = center
    c.border = border

    c = ws["E11"]
    c.value = r.fecha_reporte.strftime("%Y-%m-%d") if r.fecha_reporte else ""
    c.fill = fill_value
    c.font = normal
    c.alignment = center
    c.border = border
