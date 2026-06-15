from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from datetime import datetime
from io import BytesIO
from models.models import db, Ceco, EntregaEquipo
from routes.utils import admin_required  # ya lo tienes
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import or_

entregas_bp = Blueprint('entregas', __name__, url_prefix='/entregas')


def can_access(entrega: EntregaEquipo) -> bool:
    """Usuario normal solo puede ver lo suyo; admin ve todo."""
    return current_user.role == 'admin' or entrega.user_id == current_user.id


@entregas_bp.route('/')
@login_required
def list_entregas():
    q = request.args.get("q", "").strip()

    # Admin ve todo; user solo lo suyo
    if current_user.role == 'admin':
        query = EntregaEquipo.query
    else:
        query = EntregaEquipo.query.filter_by(user_id=current_user.id)

    # Filtro de búsqueda
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            EntregaEquipo.edificio.ilike(like),
            EntregaEquipo.piso.ilike(like),
            EntregaEquipo.area.ilike(like),
            EntregaEquipo.ceco.ilike(like),
            EntregaEquipo.jefatura.ilike(like),
            EntregaEquipo.nombre.ilike(like),
            EntregaEquipo.justificacion.ilike(like),
            EntregaEquipo.equipo_actual.ilike(like),
            EntregaEquipo.nueva_placa_sifa.ilike(like),
            EntregaEquipo.extension.ilike(like),
            EntregaEquipo.ubicacion_exacta.ilike(like),
            EntregaEquipo.observaciones.ilike(like)
        ))

    entregas = query.order_by(EntregaEquipo.id.desc()).all()
    return render_template('entregas/list.html', entregas=entregas, q=q)


@entregas_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create_entrega():
    cecos = Ceco.query.order_by(Ceco.area.asc()).all()

    if request.method == 'POST':
        ceco_id = int(request.form.get('ceco_id'))
        ceco_row = Ceco.query.get_or_404(ceco_id)

        edificio = request.form.get('edificio', '').strip()
        piso = request.form.get('piso', '').strip()
        nombre = request.form.get('nombre', '').strip()
        justificacion = request.form.get('justificacion', '').strip()
        equipo_actual = request.form.get('equipo_actual', '').strip()
        fecha_entrega_str = request.form.get('fecha_entrega', '').strip()
        observaciones = request.form.get('observaciones', '').strip()
        nueva_placa_sifa = request.form.get('nueva_placa_sifa', '').strip()
        extension = request.form.get('extension', '').strip()
        ubicacion_exacta = request.form.get('ubicacion_exacta', '').strip()

        if not (edificio and piso and nombre and justificacion and equipo_actual and fecha_entrega_str and nueva_placa_sifa):
            flash("Todos los campos obligatorios deben completarse.", "danger")
            return render_template('entregas/form.html', mode="create", cecos=cecos)

        fecha_entrega = datetime.strptime(fecha_entrega_str, "%Y-%m-%d").date()

        e = EntregaEquipo(
            user_id=current_user.id,
            ceco_id=ceco_id,

            edificio=edificio,
            piso=piso,

            area=ceco_row.area,
            ceco=ceco_row.ceco,
            jefatura=ceco_row.nombre_jefatura,

            nombre=nombre,
            justificacion=justificacion,
            equipo_actual=equipo_actual,
            fecha_entrega=fecha_entrega,
            observaciones=observaciones if observaciones else None,
            nueva_placa_sifa=nueva_placa_sifa,

            extension=extension if extension else None,
            ubicacion_exacta=ubicacion_exacta if ubicacion_exacta else None
        )

        db.session.add(e)
        db.session.commit()
        flash("Entrega creada correctamente.", "success")
        return redirect(url_for('entregas.list_entregas'))

    return render_template('entregas/form.html', mode="create", cecos=cecos)


@entregas_bp.route('/<int:entrega_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_entrega(entrega_id):
    entrega = EntregaEquipo.query.get_or_404(entrega_id)

    if not can_access(entrega):
        return "Acceso denegado", 403

    cecos = Ceco.query.order_by(Ceco.area.asc()).all()

    if request.method == 'POST':
        ceco_id = int(request.form.get('ceco_id'))
        ceco_row = Ceco.query.get_or_404(ceco_id)

        entrega.ceco_id = ceco_id
        entrega.edificio = request.form.get('edificio', '').strip()
        entrega.piso = request.form.get('piso', '').strip()

        entrega.area = ceco_row.area
        entrega.ceco = ceco_row.ceco
        entrega.jefatura = ceco_row.nombre_jefatura

        entrega.nombre = request.form.get('nombre', '').strip()
        entrega.justificacion = request.form.get('justificacion', '').strip()
        entrega.equipo_actual = request.form.get('equipo_actual', '').strip()

        fecha_entrega_str = request.form.get('fecha_entrega', '').strip()
        entrega.fecha_entrega = datetime.strptime(fecha_entrega_str, "%Y-%m-%d").date()

        obs = request.form.get('observaciones', '').strip()
        entrega.observaciones = obs if obs else None

        entrega.nueva_placa_sifa = request.form.get('nueva_placa_sifa', '').strip()

        ext = request.form.get('extension', '').strip()
        entrega.extension = ext if ext else None

        ub = request.form.get('ubicacion_exacta', '').strip()
        entrega.ubicacion_exacta = ub if ub else None

        db.session.commit()
        flash("Entrega actualizada.", "success")
        return redirect(url_for('entregas.list_entregas'))

    return render_template('entregas/form.html', mode="edit", cecos=cecos, entrega=entrega)


@entregas_bp.route('/<int:entrega_id>/delete', methods=['POST'])
@login_required
def delete_entrega(entrega_id):
    entrega = EntregaEquipo.query.get_or_404(entrega_id)

    if not can_access(entrega):
        return "Acceso denegado", 403

    db.session.delete(entrega)
    db.session.commit()
    flash("Entrega eliminada.", "info")
    return redirect(url_for('entregas.list_entregas'))


# --------- API para autollenado CECO/Jefatura al elegir Área ---------
@entregas_bp.route('/api/ceco/<int:ceco_id>')
@login_required
def api_ceco_detail(ceco_id):
    c = Ceco.query.get_or_404(ceco_id)
    return jsonify({
        "area": c.area,
        "ceco": c.ceco,
        "jefatura": c.nombre_jefatura
    })


# --------- Exportar Excel (seleccionadas o todo) ---------

@entregas_bp.route('/export', methods=['POST'])
@login_required
def export_excel():
    export_all = request.form.get('export_all') == '1'
    selected_ids = request.form.getlist('selected_ids')

    # 1) Base query según rol
    if current_user.role == 'admin':
        query = EntregaEquipo.query
    else:
        query = EntregaEquipo.query.filter_by(user_id=current_user.id)

    # 2) Filtrar seleccionadas si no es exportar todo
    if not export_all:
        if not selected_ids:
            flash("Debes seleccionar al menos una entrega para exportar.", "warning")
            return redirect(url_for('entregas.list_entregas'))
        ids_int = [int(x) for x in selected_ids]
        query = query.filter(EntregaEquipo.id.in_(ids_int))

    # ✅ AQUÍ se define rows (esto evita tu error)
    rows = query.order_by(EntregaEquipo.id.asc()).all()

    # 3) Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Entregas"

    # ✅ Encabezados EXACTOS y en el orden solicitado
    headers = [
        "Edificio",
        "Piso",
        "Area",
        "Nombre",
        "Justificación",
        "Equipo actual",
        "Fecha de entrega",
        "Observaciones",
        "Nueva placa sifa",
        "CECO",
        "Extensión",
        "Ubicación Exacta",
        "Jefatura"
    ]

    ws.append(headers)

    # ✅ Estilo encabezado azul
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # azul
    header_font = Font(bold=True, color="FFFFFF")  # blanco

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"  # congela encabezado

    # ✅ Filas en el MISMO ORDEN solicitado (SIN ID y SIN usuario)
    for e in rows:
        ws.append([
            e.edificio,
            e.piso,
            e.area,
            e.nombre,
            e.justificacion,
            e.equipo_actual,
            e.fecha_entrega,                 # como fecha real
            e.observaciones or "",
            e.nueva_placa_sifa,
            e.ceco,
            e.extension or "",
            e.ubicacion_exacta or "",
            e.jefatura
        ])

    # Formato fecha (columna 7 = G)
    for cell in ws["G"][1:]:
        cell.number_format = "yyyy-mm-dd"

    # Wrap y alineación para campos largos:
    # 5 Justificación, 8 Observaciones, 12 Ubicación Exacta
    wrap_cols = [5, 8, 12]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column in wrap_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top")

    # Anchos de columnas (para que se vea ordenado)
    widths = {
        1: 16,  # Edificio
        2: 10,  # Piso
        3: 28,  # Area
        4: 22,  # Nombre
        5: 45,  # Justificación
        6: 24,  # Equipo actual
        7: 16,  # Fecha de entrega
        8: 40,  # Observaciones
        9: 18,  # Nueva placa sifa
        10: 10, # CECO
        11: 12, # Extensión
        12: 35, # Ubicación Exacta
        13: 30  # Jefatura
    }

    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # 4) Enviar archivo
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"entregas_equipo_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

